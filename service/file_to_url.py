from concurrent.futures import ThreadPoolExecutor
from log.core.logger import get_logger
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from PIL import Image
import subprocess
import platform
import shutil
import time
import base64
import fitz
import os

logger=get_logger()

# 将图片文件转为 data URL
def image_to_data_url(image_path: str, mime_type: str = "image/png") -> str:
    with open(image_path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")
    return f"data:{mime_type};base64,{b64}"

def get_libreoffice_cmd():
    """
    自动检测 LibreOffice 命令：
    - Windows：使用手动指定路径
    - Linux/macOS：从 PATH 查找 soffice 或 libreoffice
    """
    system = platform.system().lower()

    if "windows" in system:
        # 手动指定 Windows 下 LibreOffice 路径
        win_path = r"E:\program\soffice.exe"  # 修改为你实际安装路径
        if os.path.exists(win_path):
            logger.info(f"使用 Windows LibreOffice 路径: {win_path}")
            return win_path
        else:
            logger.error(f"Windows 下未找到 LibreOffice，请检查路径: {win_path}")
            raise EnvironmentError(f"Windows 下未找到 LibreOffice，请检查路径: {win_path}")

    elif "linux" in system or "darwin" in system:
        for cmd in ["libreoffice", "soffice"]:
            if shutil.which(cmd):
                logger.info(f"使用 LibreOffice 命令: {cmd}")

                # 只在第一次运行时刷新字体缓存
                if not hasattr(get_libreoffice_cmd, "_fc_cache_done"):
                    logger.info("🔄 正在刷新字体缓存 (fc-cache -fv)...")
                    subprocess.run(["fc-cache", "-fv"], check=False)

                    # 只输出常用中文字体
                    common_cn_fonts = ["SimSun", "NSimSun", "Microsoft YaHei", "Microsoft JhengHei",
                                       "FangSong", "KaiTi", "SimHei", "WenQuanYi", "Source Han Serif", "Source Han Sans"]

                    logger.info("📋 系统可用中文字体：")
                    result = subprocess.run(["fc-list", ":family"], capture_output=True, text=True)
                    fonts = sorted(set(result.stdout.split("\n")))
                    for f in fonts:
                        if f.strip() and any(cn_font in f for cn_font in common_cn_fonts):
                            logger.info(f"  - {f.strip()}")

                    # 标记已经执行过
                    get_libreoffice_cmd._fc_cache_done = True

                return cmd

        raise EnvironmentError(
            "Linux/macOS 下未检测到 LibreOffice，请安装：\n"
            "sudo apt install libreoffice -y"
        )

    else:
        raise EnvironmentError(f"无法识别操作系统: {system}")


# 将单页 PDF 转为图片并返回 data URL
def pdf_to_image(pdf_path, page_number, output_dir="Images", dpi=100, quality=75):

    pdf = fitz.open(pdf_path)
    page = pdf[page_number]
    pix = page.get_pixmap(matrix=fitz.Matrix(dpi/72, dpi/72))

    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)
    image_path = os.path.join(output_dir, f"page_{page_number+1}.jpg")

    # 用 Pillow 保存为 JPEG
    mode = "RGBA" if pix.alpha else "RGB"
    img = Image.frombytes(mode, (pix.width, pix.height), pix.samples)
    img = img.convert("RGB")  # 确保是 JPEG 可用模式
    img.save(image_path, format="JPEG", quality=quality, optimize=True)
    logger.debug(f"已经成功转化图片: 第 {page_number + 1} 页")

    image_url = image_to_data_url(image_path, mime_type="image/jpeg")
    logger.debug(f"已经成功转化为 url: 第 {page_number + 1} 页")

    return image_url

# 并行处理 PDF 每页，返回按页码顺序的 data URL 列表
def pdf_to_url(pdf_path, max_work=10, dpi=100):
    os.makedirs("Images", exist_ok=True)

    # 获取 PDF 页数
    pdf = fitz.open(pdf_path)
    page_count = pdf.page_count
    logger.info(f"PDF共有: {page_count} 页")

    # 并行处理每页,转化成图片
    images_url = [None] * page_count
    with ThreadPoolExecutor(max_workers=max_work) as executor:
        futures = [executor.submit(pdf_to_image, pdf_path, i, "Images", dpi) for i in range(page_count)]
        for i, future in enumerate(futures):
            images_url[i] = future.result()

    return images_url, page_count


# ============================================================
# PPT → 图片
# ============================================================

def ppt_to_url(input_file: str, max_work: int, output_dir: str = "./Document"):
    start_time = time.perf_counter()


    os.makedirs(output_dir, exist_ok=True)
    libre_cmd = get_libreoffice_cmd()
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    output_pdf = os.path.join(output_dir, base_name + ".pdf")

    subprocess.run([libre_cmd, "--headless", "--convert-to", "pdf",
                    input_file, "--outdir", output_dir], check=True)

    end_time = time.perf_counter()
    elapsed = end_time - start_time
    logger.info(f"已经成功转化为pdf,耗时{elapsed:.2f}秒!")

    return pdf_to_url(output_pdf, max_work)


# ============================================================
# Word → 图片
# ============================================================

def word_to_url(input_file: str, max_work: int, output_dir: str = "./Document"):
    start_time = time.perf_counter()

    os.makedirs(output_dir, exist_ok=True)
    libre_cmd = get_libreoffice_cmd()
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    output_pdf = os.path.join(output_dir, base_name + ".pdf")

    subprocess.run([libre_cmd, "--headless", "--convert-to", "pdf",
                    input_file, "--outdir", output_dir], check=True)

    end_time = time.perf_counter()
    elapsed = end_time - start_time
    logger.info(f"已经成功转化为pdf,耗时{elapsed:.2f}秒!")

    return pdf_to_url(output_pdf, max_work)


# ============================================================
# Excel → 图片
# ============================================================
def excel_to_url(input_file: str, max_work: int, output_dir: str = "./Document"):
    start_time = time.perf_counter()

    os.makedirs(output_dir, exist_ok=True)
    libre_cmd = get_libreoffice_cmd()
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    output_pdf = os.path.join(output_dir, base_name + ".pdf")

    #将excel文件内容自动换行
    logger.info("🔧 正在格式化 Excel 文件...")
    ex = load_workbook(input_file)
    for ws in ex.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(wrap_text=True)
    ex.save(input_file)

    subprocess.run([libre_cmd, "--headless", "--convert-to", "pdf",
                    input_file, "--outdir", output_dir], check=True)

    end_time = time.perf_counter()
    elapsed = end_time - start_time
    logger.info(f"已经成功转化为pdf,耗时{elapsed:.2f}秒!")

    return pdf_to_url(output_pdf, max_work)





